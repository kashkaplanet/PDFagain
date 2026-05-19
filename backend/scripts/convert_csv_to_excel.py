import sys
import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def convert_to_excel(input_path, output_path):
    input_path = os.path.abspath(input_path)
    output_path = os.path.abspath(output_path)

    # Read CSV
    rows = []
    with open(input_path, 'r', encoding='utf-8', errors='replace') as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)

    if not rows:
        print("Error: CSV file is empty")
        sys.exit(1)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = os.path.splitext(os.path.basename(input_path))[0]

    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    # Auto-fit column widths (optional, but helpful and doesn't change data)
    for col_idx in range(len(rows[0]) if rows else 0):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
        for row in rows:
            if col_idx < len(row):
                max_length = max(max_length, len(str(row[col_idx])))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    wb.save(output_path)
    print(f"Successfully converted {input_path} to {output_path}")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python convert_csv_to_excel.py <input_csv> <output_xlsx>")
        sys.exit(1)

    convert_to_excel(sys.argv[1], sys.argv[2])
